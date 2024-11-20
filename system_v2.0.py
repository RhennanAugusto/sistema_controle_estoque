import os
import tkinter as tk
from tkinter import ttk
import datetime as dt
import sqlite3
from tkcalendar import DateEntry
import smtplib
import openpyxl
from openpyxl import Workbook
from tkinter import messagebox

banco_de_dados_existe = os.path.exists("cadastro-materiais.db")


if not banco_de_dados_existe:
    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()

    cursor.execute("""CREATE TABLE IF NOT EXISTS produtos (
        descricao TEXT,
        tipo_unidade TEXT,      
        quantidade INT,
        data_de_compra DATE,
        data_de_validade DATE,
        preco_do_produto REAL,
        empresa TEXT,
       
        data DATE  
    )""")

    conexao.commit()
    conexao.close()


entry_descrição_atualizar = None
combobox_selecionar_tipo = None
entry_quantidade_atualizar = None
entry_preço_atualizar = None
entry_data_validade_atualizar = None
entry_data_compra_atualizar = None
entry_empresa_atualizar = None
janela_atualizacao = None

lista_tipos = ["Caixa", "Saco", "Unidade", ]  



def cadastrar_materiais():
    data_criacao = dt.datetime.now()
    data_criacao = data_criacao.strftime("%d/%m/%Y %H:%M")

    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()
    cursor.execute(
        "INSERT INTO produtos (descricao, tipo_unidade, quantidade, data_de_compra, data_de_validade, preco_do_produto, empresa, data) VALUES (:descricao, :tipo_unidade, :quantidade, :data_de_compra, :data_de_validade, :preco_do_produto, :empresa, :data)",
        {
            'descricao': entry_orientacao.get(),
            'tipo_unidade': combobox_selecionar_tipo.get(),
            'quantidade': entry_quant.get(),
            'data_de_compra': entry_data.get(),
            'data_de_validade': entry_data_validade.get(),
            'preco_do_produto': entry_preco.get(),
            'empresa': entry_empresa.get(), 
            'data': data_criacao
        }
    )
    conexao.commit()
    conexao.close()

    entry_orientacao.delete(0, "end")
    combobox_selecionar_tipo.set("Selecione o tipo de unidade")
    entry_quant.delete(0, "end")
    entry_preco.delete(0, "end")
    entry_empresa.delete(0, "end")  


def carregar_itens_cadastrados():
    tree.delete(*tree.get_children())
    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()
    cursor.execute("SELECT rowid, * FROM produtos")
    itens_cadastrados = cursor.fetchall()

    data_atual = dt.datetime.now()

    for item in itens_cadastrados:

        data_validade = dt.datetime.strptime(item[5], "%d/%m/%Y")
        diferenca = data_validade - data_atual
        dias_para_vencimento = diferenca.days

        if dias_para_vencimento <= 30:
            
            tree.insert("", "end", values=item, tags=("vermelho",))
        else:
            tree.insert("", "end", values=item)


    tree.tag_configure("vermelho", background="red")

    conexao.close()
    


def atualizar_item():
    global entry_descrição_atualizar
    global entry_tipo_unidade_atualizar
    global entry_quantidade_atualizar
    global entry_data_compra_atualizar
    global entry_data_validade_atualizar
    global entry_preço_atualizar
    global entry_quantidade_venda  
    global entry_empresa_atualizar
    global janela_atualizacao

    item_selecionado = tree.selection()
    if item_selecionado:

        item = tree.item(item_selecionado)
        id_item = item["values"][0]
        janela_atualizacao = tk.Toplevel(janela)
        janela_atualizacao.title("Atualizar Item")

        label_descrição_atualizar = tk.Label(janela_atualizacao, text="Nova Descrição")
        label_descrição_atualizar.grid(column=0, row=0, padx=20, pady=10)
        entry_descrição_atualizar = tk.Entry(janela_atualizacao)
        entry_descrição_atualizar.grid(column=1, row=0, padx=20, pady=10)

        label_tipo_unidade_atualizar = tk.Label(janela_atualizacao, text="Novo Tipo de Unidade")
        label_tipo_unidade_atualizar.grid(column=0, row=1, padx=20, pady=10)
        entry_tipo_unidade_atualizar = tk.Entry(janela_atualizacao)
        entry_tipo_unidade_atualizar.grid(column=1, row=1, padx=20, pady=10)

        label_quantidade_atualizar = tk.Label(janela_atualizacao, text="Novo Lote")
        label_quantidade_atualizar.grid(column=0, row=2, padx=20, pady=10)
        entry_quantidade_atualizar = tk.Entry(janela_atualizacao)
        entry_quantidade_atualizar.grid(column=1, row=2, padx=20, pady=10)

        label_data_compra_atualizar = tk.Label(janela_atualizacao, text="Nova Data de Compra")
        label_data_compra_atualizar.grid(column=0, row=3, padx=20, pady=10)
        entry_data_compra_atualizar = tk.Entry(janela_atualizacao)
        entry_data_compra_atualizar.grid(column=1, row=3, padx=20, pady=10)

        label_data_validade_atualizar = tk.Label(janela_atualizacao, text="Nova Data de Vencimento")
        label_data_validade_atualizar.grid(column=0, row=4, padx=20, pady=10)
        entry_data_validade_atualizar = tk.Entry(janela_atualizacao)
        entry_data_validade_atualizar.grid(column=1, row=4, padx=20, pady=10)

        label_preço_atualizar = tk.Label(janela_atualizacao, text="Novo Preço do Lote")
        label_preço_atualizar.grid(column=0, row=5, padx=20, pady=10)
        entry_preço_atualizar = tk.Entry(janela_atualizacao)
        entry_preço_atualizar.grid(column=1, row=5, padx=20, pady=10)

        label_empresa_atualizar = tk.Label(janela_atualizacao, text="Nova Empresa")
        label_empresa_atualizar.grid(column=0, row=6, padx=20, pady=10)
        entry_empresa_atualizar = tk.Entry(janela_atualizacao)
        entry_empresa_atualizar.grid(column=1, row=6, padx=20, pady=10)
        
        label_quantidade_venda = tk.Label(janela_atualizacao, text="Quantidade Vendida")
        label_quantidade_venda.grid(column=0, row=7, padx=20, pady=10)
        entry_quantidade_venda = tk.Entry(janela_atualizacao)
        entry_quantidade_venda.grid(column=1, row=7, padx=20, pady=10)
        
      


        
        botao_salvar_atualização = ttk.Button(janela_atualizacao, text="Salvar Atualização",
                                              command=lambda: salvar_atualização(id_item))
        botao_salvar_atualização.grid(column=0, row=8, columnspan=2, padx=20, pady=10)

        
        entry_descrição_atualizar.insert(0, item["values"][1])
        entry_tipo_unidade_atualizar.insert(0, item["values"][2])
        entry_quantidade_atualizar.insert(0, item["values"][3])
        entry_data_compra_atualizar.insert(0, item["values"][4])
        entry_data_validade_atualizar.insert(0, item["values"][5])
        entry_preço_atualizar.insert(0, item["values"][6])
        entry_empresa_atualizar.insert(0, item["values"][7])


def salvar_atualização(id_item):
    global entry_descrição_atualizar
    global entry_tipo_unidade_atualizar
    global entry_quantidade_atualizar
    global entry_data_compra_atualizar
    global entry_data_validade_atualizar
    global entry_preço_atualizar
    global entry_quantidade_venda 
    global entry_empresa_atualizar
    global janela_atualizacao

    nova_descrição = entry_descrição_atualizar.get()
    novo_tipo_unidade = entry_tipo_unidade_atualizar.get()
    nova_quantidade = entry_quantidade_atualizar.get()
    nova_data_compra = entry_data_compra_atualizar.get()
    nova_data_validade = entry_data_validade_atualizar.get()
    novo_preço = entry_preço_atualizar.get()
    nova_empresa = entry_empresa_atualizar.get()
    quantidade_venda = entry_quantidade_venda.get()

    
    try:
        quantidade_venda = int(quantidade_venda)
    except ValueError:
        quantidade_venda = 0 

    try:
        nova_quantidade = int(nova_quantidade) - quantidade_venda
    except ValueError:
        nova_quantidade = 0  

    
    try:
        novo_preço = float(novo_preço)
    except ValueError:
        novo_preço = 0.0  

    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()

   
    if quantidade_venda > 0 and nova_quantidade <= 0:
        cursor.execute("DELETE FROM produtos WHERE rowid=?", (id_item,))
    else:
        
        cursor.execute(
            "UPDATE produtos SET descricao=?, tipo_unidade=?, quantidade=?, data_de_compra=?, data_de_validade=?, preco_do_produto=?, empresa=? WHERE rowid=?",
            (nova_descrição, novo_tipo_unidade, nova_quantidade, nova_data_compra, nova_data_validade, novo_preço, nova_empresa, id_item)
        )

    conexao.commit()
    conexao.close()
    janela_atualizacao.destroy()
    carregar_itens_cadastrados()


def atualizar_tabela_para_incluir_empresa():
    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()

    cursor.execute(""" 
    CREATE TABLE IF NOT EXISTS produtos_novo (
        descricao TEXT,
        tipo_unidade TEXT,      
        quantidade INT,
        data_de_compra DATE,
        data_de_validade DATE,
        preco_do_produto REAL,
        empresa TEXT,
        data DATE
    )
    """)

    cursor.execute("""
    INSERT INTO produtos_novo (descricao, tipo_unidade, quantidade, data_de_compra, data_de_validade, preco_do_produto, empresa, data)
    SELECT descricao, tipo_unidade, quantidade, data_de_compra, data_de_validade, preco_do_produto, '', data FROM produtos
    """)

    cursor.execute("DROP TABLE IF EXISTS produtos")

    cursor.execute("ALTER TABLE produtos_novo RENAME TO produtos")

    conexao.commit()
    conexao.close()

def verificar_e_atualizar_banco():
    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()

    cursor.execute("PRAGMA table_info(produtos)")
    colunas = [coluna[1] for coluna in cursor.fetchall()]

    if 'notificado' in colunas:
        atualizar_tabela_para_incluir_empresa()

    conexao.close()

verificar_e_atualizar_banco()


def excluir_item():
    item_selecionado = tree.selection()
    if item_selecionado:

        id_item = tree.item(item_selecionado)["values"][0]
        conexao = sqlite3.connect("cadastro-materiais.db")
        cursor = conexao.cursor()
        cursor.execute("DELETE FROM produtos WHERE rowid=?", (id_item,))
        conexao.commit()
        conexao.close()
        carregar_itens_cadastrados()  

def pesquisar_produto():
    descricao_pesquisa = entry_pesquisa.get()

    
    tree.delete(*tree.get_children())

    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()

   
    cursor.execute("SELECT rowid, * FROM produtos WHERE descricao LIKE ?", ('%' + descricao_pesquisa + '%',))
    resultados_pesquisa = cursor.fetchall()

    for resultado in resultados_pesquisa:
        tree.insert("", "end", values=resultado)

    conexao.close()
    
def exportar_para_excel():
    try:
        
        workbook = Workbook()
        planilha = workbook.active
        planilha.title = "Produtos Cadastrados"

        # Define os nomes das colunas
        colunas = ["ID", "Descrição", "Tipo de Unidade", "Quantidade", "Data de Compra", "Data de Validade", "Preço do Lote", "Empresa"]

        
        for col_num, coluna_nome in enumerate(colunas, start=1):
            planilha.cell(row=1, column=col_num, value=coluna_nome)

        
        for row_num, item in enumerate(tree.get_children(), start=2):
            valores = tree.item(item)["values"]
            for col_num, valor in enumerate(valores, start=1):
                planilha.cell(row=row_num, column=col_num, value=valor)


      
        workbook.save("produtos_cadastrados.xlsx")
        messagebox.showinfo("Exportação", "Dados exportados com sucesso para 'produtos_cadastrados.xlsx'.")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao exportar: {e}")


janela = tk.Tk()
janela.resizable(width=False, height=False)
janela.title('Cadastro de Produtos')

largura_janela = 1200
altura_janela = 745

x = (janela.winfo_screenwidth() - largura_janela) // 2
y = (janela.winfo_screenheight() - altura_janela) // 2

janela.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}")

tree = ttk.Treeview(janela, columns=("ID", "Descrição", "Tipo de Unidade", "Quantidade", "Data de Compra", "Data de Validade", "Preço do Lote", "Empresa"))

tree.grid(column=0, row=11, padx=20, pady=7, sticky="nswe", columnspan=10)

tree.column("#1", width=50, anchor="center")  # ID
tree.column("#2", width=200, anchor="w")  # Descrição
tree.column("#3", width=100, anchor="center")  # Tipo de Unidade
tree.column("#4", width=100, anchor="center")  # Quantidade
tree.column("#5", width=120, anchor="center")  # Data de Compra
tree.column("#6", width=120, anchor="center")  # Data de Validade
tree.column("#7", width=120, anchor="center")  # Preço do Lote
tree.column("#8", width=150, anchor="w")  # Empresa




tree.heading("#1", text="ID")
tree.heading("#2", text="Descrição")
tree.heading("#3", text="Tipo de Unidade")
tree.heading("#4", text="Quantidade")
tree.heading("#5", text="Data de Compra")
tree.heading("#6", text="Data de Validade")
tree.heading("#7", text="Preço do Lote")
tree.heading("#8", text="Empresa")


class EntryPlaceholder(tk.Entry):
    def __init__(self, master=None, placeholder="", color='grey', *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.placeholder = placeholder
        self.placeholder_color = color
        self.default_fg_color = self['fg']

        self.bind("<FocusIn>", self.on_entry_click)
        self.bind("<FocusOut>", self.on_focus_out)
        self.insert(0, self.placeholder)
        self['fg'] = self.placeholder_color

    def on_entry_click(self, event):
        if self.get() == self.placeholder:
            self.delete(0, "end")
            self['fg'] = self.default_fg_color

    def on_focus_out(self, event):
        if not self.get():
            self.insert(0, self.placeholder)
            self['fg'] = self.placeholder_color


label_descrição = tk.Label(janela, text="Descrição do Material")
label_descrição.grid(column=0, row=4, padx=20, pady=20, sticky="nswe", columnspan=2)

entry_pesquisa = tk.Entry(janela)
entry_pesquisa.grid(column=0, row=0, padx=20, pady=20, sticky="nswe", columnspan=2)

entry_orientacao = EntryPlaceholder(janela, placeholder="Digite a descrição do produto")
entry_orientacao.grid(column=0, row=4, padx=20, pady=20, sticky="nswe", columnspan=10)

label_unidade = tk.Label(janela, text="Tipo de Unidade do Material")
label_unidade.grid(column=0, row=5, padx=20, pady=20, sticky="nswe", columnspan=2)

combobox_selecionar_tipo = ttk.Combobox(janela, values=lista_tipos, state='readonly')
combobox_selecionar_tipo.set("Selecione o tipo de unidade")
combobox_selecionar_tipo.grid(column=0, row=5, padx=20, pady=20, sticky="nswe", columnspan=10)

label_quant = tk.Label(janela, text="Quantidade do Material")
label_quant.grid(column=0, row=6, padx=20, pady=20, sticky='nswe', columnspan=10)

entry_quant = EntryPlaceholder(janela, placeholder="Digite a quantidade do produto")
entry_quant.grid(column=0, row=6, padx=20, pady=20, sticky="nswe", columnspan=10)

label_data = tk.Label(janela)
label_data.grid(column=0, row=7, padx=20, pady=20, sticky='nswe', columnspan=1)

entry_data = DateEntry(janela, date_pattern="dd/mm/yyyy")
entry_data.grid(column=1, row=7, padx=20, pady=20, sticky="nswe", columnspan=4)

label_placeholder_compra = tk.Label(janela, text="Data de Compra do Produto")
label_placeholder_compra.grid(column=0, row=7, padx=5, pady=20, sticky='e', columnspan=1)

label_data_validade = tk.Label(janela)
label_data_validade.grid(column=5, row=7, padx=20, pady=20, sticky='nswe', columnspan=1)

entry_data_validade = DateEntry(janela, date_pattern="dd/mm/yyyy")
entry_data_validade.grid(column=6, row=7, padx=20, pady=20, sticky="nswe", columnspan=4)

label_placeholder_vencimento = tk.Label(janela, text="Data de Vencimento do Produto")
label_placeholder_vencimento.grid(column=5, row=7, padx=5, pady=20, sticky='e', columnspan=1)

label_preco = tk.Label(janela, text="Preço do Lote")
label_preco.grid(column=0, row=9, padx=20, pady=20, sticky='nswe', columnspan=10)

entry_preco = EntryPlaceholder(janela, placeholder="Digite o preço do lote")
entry_preco.grid(column=0, row=9, padx=20, pady=20, sticky="nswe", columnspan=10)

# Adição do campo Nome da Empresa ao lado do campo de pesquisa
label_empresa = tk.Label(janela, text="Nome da Empresa")
label_empresa.grid(column=2, row=0, padx=20, pady=20, sticky="nswe", columnspan=1)  # Ao lado do campo de pesquisa

entry_empresa = EntryPlaceholder(janela, placeholder="Digite o nome da empresa")
entry_empresa.grid(column=3, row=0, padx=20, pady=20, sticky="nswe", columnspan=1)  # Ao lado do campo de pesquisa

botao_cadastrar_produtos = tk.Button(text='Cadastrar Produtos', command=cadastrar_materiais)
botao_cadastrar_produtos.grid(column=0, row=10, padx=20, pady=20, sticky="nswe", columnspan=10)

botao_atualizar = ttk.Button(text="Atualizar Item", command=atualizar_item)
botao_atualizar.grid(column=0, row=12, padx=20, pady=20, columnspan=7)

botao_exportar = ttk.Button(janela, text="Exportar para Excel", command=exportar_para_excel)
botao_exportar.grid(column=1, row=12, padx=20, pady=10, columnspan=8)

botao_excluir = ttk.Button(text="Excluir Item", command=excluir_item)
botao_excluir.grid(column=3, row=12, padx=20, pady=20, columnspan=9)

botao_pesquisar = tk.Button(text='Pesquisar', command=pesquisar_produto)
botao_pesquisar.grid(column=2, row=0, padx=10, pady=20, sticky="nswe", columnspan=1)

carregar_itens_cadastrados()

janela.mainloop()