import os
import tkinter as tk
from tkinter import ttk
import datetime as dt
import sqlite3
from tkcalendar import DateEntry
import smtplib
import openpyxl
from openpyxl import Workbook
from tkinter import messagebox

banco_de_dados_existe = os.path.exists("cadastro-materiais.db")

# Se o banco de dados não existe, vai ser criado
if not banco_de_dados_existe:
    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()

    cursor.execute("""CREATE TABLE IF NOT EXISTS produtos (
        descricao TEXT,
        tipo_unidade TEXT,      
        quantidade INT,
        data_de_compra DATE,
        data_de_validade DATE,
        preco_do_produto REAL,
        notificado INT DEFAULT 0,
        data DATE  
    )""")

    conexao.commit()
    conexao.close()

# vou definir as variáveis globais para os widgets de atualização
entry_descrição_atualizar = None
combobox_selecionar_tipo = None
entry_quantidade_atualizar = None
entry_preço_atualizar = None
entry_data_validade_atualizar = None
entry_data_compra_atualizar = None
janela_atualizacao = None

lista_tipos = ["Caixa", "Saco", "Unidade", ]  # lista para que nosso combobox consiga adicionar esse valores
def enviar_aviso_por_email(descricao, data_vencimento):

    servidor_smtp = 'smtp.gmail.com'  # servidor SMTP do meu de e-mail
    porta_smtp = 587
    email_origem = 'rhennanaugusto23@gmail.com'
    senha = 'vrbq beto tpde tdoc'


    server = smtplib.SMTP(servidor_smtp, porta_smtp)
    server.starttls()
    server.login(email_origem, senha)


    assunto = 'Aviso de Vencimento de Produto'

    mensagem = f'O produto "{descricao}" está a 30 dias ou menos do vencimento em {data_vencimento}.'
    mensagem = f'Subject: {assunto}\n\n{mensagem}'


    destinatarios = ['augustimparzival@gmail.com']


    for destinatario in destinatarios:
        server.sendmail(email_origem, destinatario, mensagem.encode('utf-8'))


    server.quit()
def verificar_e_atualizar_notificacoes():
    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()
    cursor.execute("SELECT rowid, * FROM produtos WHERE notificado=0")
    itens_notificar = cursor.fetchall()

    # Obtém a data atual
    data_atual = dt.datetime.now()

    for item in itens_notificar:

        data_validade = dt.datetime.strptime(item[5], "%d/%m/%Y")
        diferenca = data_validade - data_atual
        dias_para_vencimento = diferenca.days

        if dias_para_vencimento <= 30:
            # Envia um aviso por e-mail
            enviar_aviso_por_email(item[1], item[5])

            cursor.execute("UPDATE produtos SET notificado=1 WHERE rowid=?", (item[0],)) # aqui ta att a coluna de notificado

    conexao.commit()
    conexao.close()
def cadastrar_materiais():
    data_criacao = dt.datetime.now()
    data_criacao = data_criacao.strftime("%d/%m/%Y %H:%M")

    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()
    cursor.execute(
        "INSERT INTO produtos VALUES(:descricao, :tipo_unidade, :quantidade, :data_de_compra, :data_de_validade, :preco_do_produto, :notificado, :data)",
        {
            'descricao': entry_orientacao.get(),
            'tipo_unidade': combobox_selecionar_tipo.get(),
            'quantidade': entry_quant.get(),
            'data_de_compra': entry_data.get(),
            'data_de_validade': entry_data_validade.get(),
            'preco_do_produto': entry_preco.get(),
            'notificado': 0,
            'data': data_criacao
        }
    )
    conexao.commit()
    conexao.close()

    entry_orientacao.delete(0, "end")
    combobox_selecionar_tipo.set("Selecione o tipo de unidade")
    entry_quant.delete(0, "end")
    entry_preco.delete(0, "end")
def carregar_itens_cadastrados():
    tree.delete(*tree.get_children())
    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()
    cursor.execute("SELECT rowid, * FROM produtos")
    itens_cadastrados = cursor.fetchall()

    data_atual = dt.datetime.now()

    for item in itens_cadastrados:

        data_validade = dt.datetime.strptime(item[5], "%d/%m/%Y")
        diferenca = data_validade - data_atual
        dias_para_vencimento = diferenca.days

        if dias_para_vencimento <= 30:
            # definir a cor vermelha para indicar que venceu
            tree.insert("", "end", values=item, tags=("vermelho",))
        else:
            tree.insert("", "end", values=item)


    tree.tag_configure("vermelho", background="red")

    conexao.close()
    verificar_e_atualizar_notificacoes()


def atualizar_item():
    global entry_descrição_atualizar
    global entry_tipo_unidade_atualizar
    global entry_quantidade_atualizar
    global entry_data_compra_atualizar
    global entry_data_validade_atualizar
    global entry_preço_atualizar
    global janela_atualizacao

    item_selecionado = tree.selection()
    if item_selecionado:

        item = tree.item(item_selecionado)
        id_item = item["values"][0]
        janela_atualizacao = tk.Toplevel(janela)
        janela_atualizacao.title("Atualizar Item")


        label_descrição_atualizar = tk.Label(janela_atualizacao, text="Nova Descrição")
        label_descrição_atualizar.grid(column=0, row=0, padx=20, pady=10)
        entry_descrição_atualizar = tk.Entry(janela_atualizacao)
        entry_descrição_atualizar.grid(column=1, row=0, padx=20, pady=10)

        label_tipo_unidade_atualizar = tk.Label(janela_atualizacao, text="Novo Tipo de Unidade")
        label_tipo_unidade_atualizar.grid(column=0, row=1, padx=20, pady=10)
        entry_tipo_unidade_atualizar = tk.Entry(janela_atualizacao)
        entry_tipo_unidade_atualizar.grid(column=1, row=1, padx=20, pady=10)

        label_quantidade_atualizar = tk.Label(janela_atualizacao, text="Novo Lote")
        label_quantidade_atualizar.grid(column=0, row=2, padx=20, pady=10)
        entry_quantidade_atualizar = tk.Entry(janela_atualizacao)
        entry_quantidade_atualizar.grid(column=1, row=2, padx=20, pady=10)

        label_data_compra_atualizar = tk.Label(janela_atualizacao, text="Nova Data de Compra")
        label_data_compra_atualizar.grid(column=0, row=3, padx=20, pady=10)
        entry_data_compra_atualizar = tk.Entry(janela_atualizacao)
        entry_data_compra_atualizar.grid(column=1, row=3, padx=20, pady=10)

        label_data_validade_atualizar = tk.Label(janela_atualizacao, text="Nova Data de Vencimento")
        label_data_validade_atualizar.grid(column=0, row=4, padx=20, pady=10)
        entry_data_validade_atualizar = tk.Entry(janela_atualizacao)
        entry_data_validade_atualizar.grid(column=1, row=4, padx=20, pady=10)

        label_preço_atualizar = tk.Label(janela_atualizacao, text="Novo Preço do Lote")
        label_preço_atualizar.grid(column=0, row=5, padx=20, pady=10)
        entry_preço_atualizar = tk.Entry(janela_atualizacao)
        entry_preço_atualizar.grid(column=1, row=5, padx=20, pady=10)

        # Botão para salvar a atualização
        botao_salvar_atualização = ttk.Button(janela_atualizacao, text="Salvar Atualização",
                                              command=lambda: salvar_atualização(id_item))
        botao_salvar_atualização.grid(column=0, row=6, columnspan=2, padx=20, pady=10)


        entry_descrição_atualizar.insert(0, item["values"][1])
        entry_tipo_unidade_atualizar.insert(0, item["values"][2])
        entry_quantidade_atualizar.insert(0, item["values"][3])
        entry_data_compra_atualizar.insert(0, item["values"][4])
        entry_data_validade_atualizar.insert(0, item["values"][5])
        entry_preço_atualizar.insert(0, item["values"][6])



def salvar_atualização(id_item):
    global entry_descrição_atualizar
    global entry_tipo_unidade_atualizar
    global entry_quantidade_atualizar
    global entry_data_compra_atualizar
    global entry_data_validade_atualizar
    global entry_preço_atualizar
    global janela_atualizacao

    nova_descrição = entry_descrição_atualizar.get()
    novo_tipo_unidade = entry_tipo_unidade_atualizar.get()
    nova_quantidade = entry_quantidade_atualizar.get()
    nova_data_compra = entry_data_compra_atualizar.get()
    nova_data_validade = entry_data_validade_atualizar.get()
    novo_preço = entry_preço_atualizar.get()


    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()
    cursor.execute(
        "UPDATE produtos SET descricao=?, tipo_unidade=?, quantidade=?, data_de_compra=?, data_de_validade=?, preco_do_produto=? WHERE rowid=?",
        (nova_descrição, novo_tipo_unidade, nova_quantidade, nova_data_compra, nova_data_validade, novo_preço, id_item))
    conexao.commit()
    conexao.close()
    janela_atualizacao.destroy()
    carregar_itens_cadastrados()


def excluir_item():
    item_selecionado = tree.selection()
    if item_selecionado:

        id_item = tree.item(item_selecionado)["values"][0]
        conexao = sqlite3.connect("cadastro-materiais.db")
        cursor = conexao.cursor()
        cursor.execute("DELETE FROM produtos WHERE rowid=?", (id_item,))
        conexao.commit()
        conexao.close()
        carregar_itens_cadastrados()  # Atualiza a TreeView após a exclusão

def pesquisar_produto():
    descricao_pesquisa = entry_pesquisa.get()

    # Limpa a TreeView antes de exibir os resultados da pesquisa
    tree.delete(*tree.get_children())

    conexao = sqlite3.connect("cadastro-materiais.db")
    cursor = conexao.cursor()

    # Realiza a pesquisa no banco de dados
    cursor.execute("SELECT rowid, * FROM produtos WHERE descricao LIKE ?", ('%' + descricao_pesquisa + '%',))
    resultados_pesquisa = cursor.fetchall()

    for resultado in resultados_pesquisa:
        tree.insert("", "end", values=resultado)

    conexao.close()
    
def exportar_para_excel():
    try:
        # Cria um novo workbook e seleciona a planilha ativa
        workbook = Workbook()
        planilha = workbook.active
        planilha.title = "Produtos Cadastrados"

        # Define os nomes das colunas
        colunas = ["ID", "Descrição", "Tipo de Unidade", "Quantidade", "Data de Compra", "Data de Validade", "Preço do Lote"]

        # Adiciona os nomes das colunas na primeira linha do Excel
        for col_num, coluna_nome in enumerate(colunas, start=1):
            planilha.cell(row=1, column=col_num, value=coluna_nome)

        # Adiciona os dados da TreeView no arquivo Excel
        for row_num, item in enumerate(tree.get_children(), start=2):
            valores = tree.item(item)["values"]
            for col_num, valor in enumerate(valores, start=1):
                planilha.cell(row=row_num, column=col_num, value=valor)

        # Salva o arquivo Excel
        workbook.save("produtos_cadastrados.xlsx")
        messagebox.showinfo("Exportação", "Dados exportados com sucesso para 'produtos_cadastrados.xlsx'.")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao exportar: {e}")


janela = tk.Tk()
janela.resizable(width=False, height=False)
janela.title('Cadastro de Produtos')

largura_janela = 1050
altura_janela = 745

x = (janela.winfo_screenwidth() - largura_janela) // 2
y = (janela.winfo_screenheight() - altura_janela) // 2

janela.geometry(f"{largura_janela}x{altura_janela}+{x}+{y}")

tree = ttk.Treeview(janela, columns=(
    "ID", "Descrição", "Tipo de Unidade", "Quantidade", "Data de Compra", "Data de Validade", "Preço do Lote"))
tree.grid(column=0, row=11, padx=20, pady=7, sticky="nswe", columnspan=10)

tree.column("#1", width=50, anchor="center")  # ID
tree.column("#2", width=200, anchor="w")  # Descrição
tree.column("#3", width=100, anchor="center")  # Tipo de Unidade
tree.column("#4", width=100, anchor="center")  # Quantidade
tree.column("#5", width=120, anchor="center")  # Data de Compra
tree.column("#6", width=120, anchor="center")  # Data de Validade
tree.column("#7", width=120, anchor="center")  # Preço do Lote

# Define os cabeçalhos das colunas
tree.heading("#1", text="ID")
tree.heading("#2", text="Descrição")
tree.heading("#3", text="Tipo de Unidade")
tree.heading("#4", text="Quantidade")
tree.heading("#5", text="Data de Compra")
tree.heading("#6", text="Data de Validade")
tree.heading("#7", text="Preço do Lote")


class EntryPlaceholder(tk.Entry):
    def __init__(self, master=None, placeholder="", color='grey', *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.placeholder = placeholder
        self.placeholder_color = color
        self.default_fg_color = self['fg']

        self.bind("<FocusIn>", self.on_entry_click)
        self.bind("<FocusOut>", self.on_focus_out)
        self.insert(0, self.placeholder)
        self['fg'] = self.placeholder_color

    def on_entry_click(self, event):
        if self.get() == self.placeholder:
            self.delete(0, "end")
            self['fg'] = self.default_fg_color

    def on_focus_out(self, event):
        if not self.get():
            self.insert(0, self.placeholder)
            self['fg'] = self.placeholder_color


label_descrição = tk.Label(janela, text="Descrição do Material")
label_descrição.grid(column=0, row=4, padx=20, pady=20, sticky="nswe", columnspan=2)
entry_pesquisa = tk.Entry(janela)
entry_pesquisa.grid(column=0, row=0, padx=20, pady=20, sticky="nswe", columnspan=2)
entry_orientacao = EntryPlaceholder(janela, placeholder="Digite a descrição do produto")
entry_orientacao.grid(column=0, row=4, padx=20, pady=20, sticky="nswe", columnspan=10)
label_unidade = tk.Label(janela, text="Tipo de Unidade do Material")
label_unidade.grid(column=0, row=5, padx=20, pady=20, sticky="nswe", columnspan=2)
combobox_selecionar_tipo = ttk.Combobox(janela, values=lista_tipos, state='readonly')
combobox_selecionar_tipo.set("Selecione o tipo de unidade")
combobox_selecionar_tipo.grid(column=0, row=5, padx=20, pady=20, sticky="nswe", columnspan=10)
label_quant = tk.Label(janela, text="Quantidade do Material")
label_quant.grid(column=0, row=6, padx=20, pady=20, sticky='nswe', columnspan=10)
entry_quant = EntryPlaceholder(janela, placeholder="Digite a quantidade do produto")
entry_quant.grid(column=0, row=6, padx=20, pady=20, sticky="nswe", columnspan=10)
label_data = tk.Label(janela)
label_data.grid(column=0, row=7, padx=20, pady=20, sticky='nswe', columnspan=1)
entry_data = DateEntry(janela, date_pattern="dd/mm/yyyy")
entry_data.grid(column=1, row=7, padx=20, pady=20, sticky="nswe", columnspan=4)
label_placeholder_compra = tk.Label(janela, text="Data de Compra do Produto")
label_placeholder_compra.grid(column=0, row=7, padx=5, pady=20, sticky='e', columnspan=1)
label_data_validade = tk.Label(janela)
label_data_validade.grid(column=5, row=7, padx=20, pady=20, sticky='nswe', columnspan=1)
entry_data_validade = DateEntry(janela, date_pattern="dd/mm/yyyy")
entry_data_validade.grid(column=6, row=7, padx=20, pady=20, sticky="nswe", columnspan=4)
label_placeholder_vencimento = tk.Label(janela, text="Data de Vencimento do Produto")
label_placeholder_vencimento.grid(column=5, row=7, padx=5, pady=20, sticky='e', columnspan=1)
label_preco = tk.Label(janela, text="Preço do Lote")
label_preco.grid(column=0, row=9, padx=20, pady=20, sticky='nswe', columnspan=10)
entry_preco = EntryPlaceholder(janela, placeholder="Digite o preço do lote")
entry_preco.grid(column=0, row=9, padx=20, pady=20, sticky="nswe", columnspan=10)
botao_cadastrar_produtos = tk.Button(text='Cadastrar Produtos', command= cadastrar_materiais)
botao_cadastrar_produtos.grid(column=0, row=10, padx=20, pady=20, sticky="nswe", columnspan=10)
botao_atualizar = ttk.Button(text="Atualizar Item", command= atualizar_item)
botao_atualizar.grid(column=0, row=12, padx=20, pady=20,  columnspan=7)
botao_exportar = ttk.Button(janela, text="Exportar para Excel", command=exportar_para_excel)
botao_exportar.grid(column=1, row=12, padx=20, pady=10, columnspan=8)
botao_excluir = ttk.Button(text="Excluir Item", command= excluir_item)
botao_excluir.grid(column=3, row=12, padx=20, pady=20,  columnspan=9)
botao_pesquisar = tk.Button(text='Pesquisar', command=pesquisar_produto)
botao_pesquisar.grid(column=2, row=0, padx=10, pady=20, sticky="nswe", columnspan=1)

# da linha 44 ate a linha 65 é toda a parte grafica e visual do nosso cadastro

carregar_itens_cadastrados()

janela.mainloop()